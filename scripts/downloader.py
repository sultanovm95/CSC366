import requests
from bs4 import BeautifulSoup
import wget
import pprint
import ssl
from os.path import exists
from multiprocessing import Pool

ssl._create_default_https_context = ssl._create_unverified_context


def download_csv(criteria, url, path="../data/criteria"):
    filename = url.split("/")[-1]
    filename = filename.replace("?fmt=csv", "")

    if exists(f"{path}/{filename}"):
        print(f"{filename} exists")
        return
    print(url)

    page = requests.get(url)
    with open(f"{path}/+{criteria}+{filename}", "wb") as f:
        f.write(page.content)


def get_url(page_url):
    criteria, url = page_url

    page = requests.get(url)
    soup = BeautifulSoup(page.content, "html.parser")
    csv_url = soup.find({"h2": {"class": "reportdesc"}}).find_all(
        {"a": {"class": "ms-2"}}
    )[1]["href"]

    csv_url = "https://www.onetonline.org" + csv_url
    print(criteria, csv_url)
    download_csv(criteria, csv_url)


def retrieve_links(filename: str):
    import openpyxl

    urls = []
    wb = openpyxl.load_workbook(filename)
    ws = wb["Sheet1"]

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            criteria = ws.cell(row=r, column=1).value
            hyperlink = ws.cell(row=r, column=c).hyperlink
            if hyperlink is not None and criteria is not None:
                urls.append((criteria, hyperlink.target))
    pprint.pprint(urls)
    return urls


if __name__ == "__main__":
    urls = retrieve_links(
        "../data/info/O-Net Computer Science Template for Survey_5_13_2022.xlsx"
    )

    print(f"Number of URLs: {len(urls)}")

    with Pool(10) as p:
        p.map(get_url, urls)
